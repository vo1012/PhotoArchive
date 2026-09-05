"""Детализация прогонного отчёта — единый xlsx (PROMPT_report_detail_xlsx.md, Фаза 0/1,
реализация начата 2026-08-16). Заменяет собой план трёх HTML companion-страниц из
PROMPT_report_run_redesign.md -- одна плоская таблица (одна строка = одно событие source ->
dest/дуп/не прочитано), фильтрация/сортировка средствами самого Excel (автофильтр, сортировка
по колонке), не жёстко закодированная HTML-структура. (До 2026-08-28 поверх был ещё Excel-
outline — сворачиваемые группы по папке; убран вместе с переходом на write_only-режим openpyxl,
см. _write_flat_xlsx().)

Отдельный модуль (не report.py) -- по решению пользователя 2026-08-16 ("сам" реши, разбивать
ли исходники): report.py уже перевалил за 5000 строк, этот файл -- отдельная, самодостаточная
ответственность (сборка плоских строк + запись .xlsx), не переиспользуется HTML-рендером
report.py вообще. Импортирует несколько мелких чистых хелперов ИЗ report.py (одностороннее
направление, report.py импортирует этот модуль только локально внутри функции -- см.
generate_report()/_generate_from_model() -- иначе был бы циклический импорт на уровне модуля).

Область действия (см. спеку, "Формат и охват"): оба режима, где уже существует ветка
checklist_new is not None в _generate_from_model() -- реальный прогон (level=="target") и
dry-run/интерактивный предпросмотр (level=="workdir", после унификации 2026-08-14 -- один
код-путь с level=="target", разница только в том, физически ли создан dest). Данные -- ТОЛЬКО
data_new (_split_rows_by_time(), см. generate_report()) -- этот xlsx только про текущий прогон,
не история архива целиком (см. спеку, "Инварианты").

Паспорт архива (PROMPT_report_detail_xlsx.md, "Открыто на момент записи", решено и реализовано
2026-08-16) -- ВТОРОЙ, отдельный построитель/генератор в этом же модуле (_build_passport_detail_
rows()/generate_passport_detail_xlsx()): self-scan TARGET, не source->dest события прогона --
другая форма строки (находка, не файл-событие), общая механика записи листа (_write_flat_xlsx()),
визуальные хелперы (_argb()/цвета) и с 2026-09-05 -- PowerShell-хелперы (_powershell_open_
command()/_series_ps_commands()) для одноимённого столбца, теперь в обоих построителях."""
import os
import re

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from report import (
    CATEGORY_PALETTE,
    COLOR_ACCENT_SECONDARY,
    COLOR_TEXT_MUTED,
    _cluster_near_dup,
    _cluster_passport_edges,
    _dispute_reason_label,
    _ext,
    _media_kind,
    _n_files,
    _source_basename,
    _winlong,
)

DETAIL_XLSX_FILENAME = "report_detail.xlsx"

_COLUMN_HEADERS = [
    "Путь к исходной папке", "Имя файла", "Расширение", "Тип медиа", "Копировано",
    "Куда / с чем дуп", "Итоговое имя файла", "№ серии", "Примечание",
    "Открыть файл (PowerShell)",
]
_COLUMN_WIDTHS = [55, 32, 10, 10, 10, 55, 32, 8, 45, 46]
# 0-based -- индекс "Открыть файл (PowerShell)" в _values ниже. ПОСЛЕДНЯЯ колонка -- живая
# находка пользователя, 2026-09-05: длинный текст команды (теперь -- вся серия, не один файл,
# см. _series_ps_commands()) переполняет соседнюю пустую ячейку Excel'ем визуально; последняя
# колонка переполняется в пустое место листа, не на "Примечание".
_DETAIL_PS_COMMAND_COL = 9

_KIND_LABELS = {"image": "фото", "video": "видео", "raw": "RAW", "other": "прочее"}

# Раскраска -- решение пользователя 2026-08-15 (спека, "Визуальное различение"): только
# вспомогательный сигнал, категория строки по-прежнему однозначно читается из
# "Копировано"+"Примечание". Обычные скопированные/скопированные-с-пометкой -- без
# стилизации (None ниже, читается как чёрный по умолчанию).
_COLOR_DUPLICATE = COLOR_TEXT_MUTED  # тот же серый, что уже задаёт .muted в HTML-отчёте
_COLOR_RAW_SKIPPED = CATEGORY_PALETTE[3]  # свой цвет, отдельный от серого/терракотового
_COLOR_PROBLEM = COLOR_ACCENT_SECONDARY  # "использовать скупо" -- только unreadable.csv


def _argb(hex6: str) -> str:
    """openpyxl.styles.Font(color=...) молча трактует голый 6-значный RGB как ARGB с
    alpha=00 (полностью прозрачный) вместо ожидаемого непрозрачного цвета -- проверено
    исполнением (round-trip через openpyxl.load_workbook() при реализации этого модуля).
    Явный alpha=FF -- единственный надёжный способ получить видимый цвет."""
    return "FF" + hex6.lstrip("#").upper()


def _source_dirname(path: str) -> str:
    """Как _win_dirname() в report.py, но понимает и "/" -- origin_display для файлов ИЗ
    АРХИВА использует "/" внутри member-пути (см. _source_basename() в report.py, тот же
    повод), report.py-шная _win_dirname() рассчитана только на dest/matched-пути (всегда
    "\\", см. её докстринг) -- для source-путей этого модуля нужен свой вариант."""
    m = re.search(r"^(.*)[\\/][^\\/]*$", path or "")
    return m.group(1) if m else ""


_VIDEO_TS_DIR_RE = re.compile(r"^VIDEO_TS( \(\d+\))?$")


def _dvd_unit_root(dest: str) -> str:
    """Раунд 96 (придирка, закрыта раунд 116-раунд ответа): группировка DVD-строк раньше брала
    просто os.path.dirname(dest) -- для нестандартного рипа с файлами во вложенной подпапке
    ВНУТРИ VIDEO_TS (_dvd_unit_file_records(), photosort_win.py, явно рекурсивна "на случай
    нестандартного рипа") это раскалывало один физический юнит на несколько групп/строк, ровно
    тот класс проблемы, что уже был находкой Раунда 95, просто для более узкого случая. Вместо
    ближайшего родителя -- ищем ближайшего ПРЕДКА с именем "VIDEO_TS"/"VIDEO_TS (N)" (буквальное
    имя папки-юнита, см. _unique_dvd_dest_name(), base_name="VIDEO_TS"); если такого предка нет
    (не должно случаться при штатной сборке DVD-юнита) -- откат на прежнее поведение
    (dirname(dest)), не падаем."""
    current = dest or ""
    while current:
        if _VIDEO_TS_DIR_RE.match(_source_basename(current)):
            return current
        parent = _source_dirname(current)
        if parent == current:
            break
        current = parent
    return _source_dirname(dest)


def _powershell_open_command(path: str) -> str:
    """SESSION-HANDOFF.txt, 2026-09-04 ("серии: РЕШЕНИЕ ПРИНЯТО"): готовая команда для вставки
    в открытое окно PowerShell -- открывает файл приложением по умолчанию для его расширения у
    ЭТОГО пользователя (не хардкодим "Фотографии"), тем же путём, что уже проверен живым кликом
    в этой сессии (ShellExecute без явной программы). Одинарная кавычка в пути ломает
    PowerShell-литерал '...' -- экранируется удвоением, штатный приём самого PowerShell."""
    return "Start-Process '%s'" % path.replace("'", "''")


# REVIEW-HANDOFF.md, Раунд 207 [ЗАМЕЧАНИЕ] 207-1: запас от жёсткого лимита ячейки .xlsx
# (32767 UTF-16 code units, формат, не openpyxl-специфика) -- без запаса openpyxl/Excel
# молча обрезают значение при реальной записи, обрезка может прийтись на середину
# 'литерала...' (нечётное число кавычек) -- ВСЯ команда становится невыполнимой в PowerShell
# целиком ("The string is missing the terminator"), не "открылись первые N файлов". Проверено
# исполнением: реалистичный кластер серийной съёмки (~350-400+ членов при типичной длине
# пути) на большом архиве даёт команду длиннее лимита. 30000 -- запас с большим отрывом от
# 32767 под замыкающий комментарий-пояснение ниже (тот всегда короче нескольких сотен
# символов).
_PS_COMMAND_CHAR_BUDGET = 30000


def _series_ps_commands(abs_paths: list) -> dict:
    """Одна команда на ВЕСЬ кластер (серию/группу дублей) -- открывает КАЖДЫЙ файл серии
    сразу, независимо от того, лежат ли они в одной папке. Решение пользователя, 2026-09-05:
    отменяет прежнее ограничение "команда только если >=2 членов в одной папке" (2026-09-04) --
    та причина опиралась на то, что Проводник открывается по одной папке за раз, но эта колонка
    Проводник вообще не открывает: Start-Process запускает каждый файл СВОЕЙ программой по
    умолчанию, расположение файла для визуального сравнения роли не играет. ВСЕ члены кластера
    получают ОДНУ и ТУ ЖЕ строку (копипаст с любой строки серии открывает всю серию целиком) --
    пути сортируются для детерминированности (один и тот же кластер даёт один и тот же текст
    команды при каждом прогоне, порядок членов не зависит от порядка обхода). Общая для
    Паспорта (кластеры self-scan) и обычного прогона (кластеры near_dup внутри
    _build_detail_rows) -- одна и та же форма входа (список абсолютных путей одного кластера,
    гарантированно >=2 элемента -- кластер существует только при наличии хотя бы одного ребра).

    REVIEW-HANDOFF.md, Раунд 207 (207-1): гигантский кластер обрезается по ЦЕЛЫМ путям (никогда
    не разрывает Start-Process '...'-литерал посередине) под _PS_COMMAND_CHAR_BUDGET, остаток
    обозначается PowerShell-КОММЕНТАРИЕМ (после `#` до конца строки) -- сама команда остаётся
    синтаксически рабочей, пользователь явно видит текстом, что кластер не поместился целиком,
    вместо молчаливого битого обрезка от Excel.

    REVIEW-HANDOFF.md, Раунд 208 (208-1, придирка на 43d4062): упаковка ПРОПУСКАЕТ (`continue`)
    отдельный не поместившийся элемент, а не останавливает набор целиком (`break`) -- один
    аномально длинный путь в алфавитно РАННЕЙ позиции (сам по себе или с уже накопленным не
    помещается в бюджет) раньше отбрасывал вообще ВСЕ элементы после себя, даже если почти
    весь кластер состоял из коротких, прекрасно помещающихся путей. С `continue` результат --
    максимум того, что реально влезает (пропуская только сами непомещающиеся элементы), не
    "префикс алфавитного порядка до первого сбоя"."""
    ordered = sorted(abs_paths)
    pieces = [_powershell_open_command(p) for p in ordered]
    command = "; ".join(pieces)
    if len(command) > _PS_COMMAND_CHAR_BUDGET:
        fitted = []
        length = 0
        for piece in pieces:
            add = len(piece) + (2 if fitted else 0)  # "; " перед каждым, кроме первого
            if length + add > _PS_COMMAND_CHAR_BUDGET:
                continue  # пропускаем ТОЛЬКО этот элемент, не рвём набор остальных
            fitted.append(piece)
            length += add
        omitted = len(pieces) - len(fitted)
        note = f"# и ещё {omitted} файлов не поместились в команду (слишком большой кластер)"
        command = f"{'; '.join(fitted)}; {note}" if fitted else note
    return dict.fromkeys(abs_paths, command)


def _dup_pair_ps_command(source: str, matched: str) -> str:
    """Дубликат обычного прогона -- не кластер self-scan (Паспорт), а ровно ОДНА пара:
    новый файл (ещё не скопирован, лежит на SOURCE) vs то, с чем он совпал (уже в архиве,
    TARGET). Команда открывает ОБА -- тот же принцип "открыть всю группу сразу", что и у
    _series_ps_commands() выше, просто без сортировки/join по кластеру: группа здесь всегда
    ровно из двух файлов, порядок (источник, потом совпадение) сам по себе осмыслен."""
    if not source or not matched:
        return ""
    return _powershell_open_command(source) + "; " + _powershell_open_command(matched)


def _row_kind(primary_path: str, fallback_path: str = "") -> str:
    """image/video/raw -- три основных значения колонки "Тип медиа" (спека). "other" --
    осознанный запасной вариант ДЛЯ ЭТОГО модуля, не буквально "ровно три значения" из
    спеки: у спорных файлов (media_note="not_media") расширение может не входить ни в один
    распознаваемый набор вообще -- см. докстринг generate_detail_xlsx() ниже, где это явно
    объясняется как сознательное отступление, а не молчаливая недоработка."""
    kind = _media_kind(primary_path)
    if kind == "other" and fallback_path:
        kind = _media_kind(fallback_path)
    return kind


def _build_detail_rows(data: dict) -> list:
    """Плоский построитель "с нуля" -- PROMPT_report_detail_xlsx.md, раздел "Данные —
    построитель": готового построителя такого уровня в report.py нет (там только агрегаты
    сводного экрана, build_model_from_rows()), near_dup_clusters -- единственная часть,
    переиспользуемая как есть (_cluster_near_dup(), тот же граф, что уже питает Раздел 3.3
    сводного экрана).

    Возвращает список dict (не строки openpyxl напрямую) -- сборка данных отделена от записи
    в xlsx, тестируется независимо от openpyxl."""
    near_dup_clusters = _cluster_near_dup(data.get("near_dup_edges", []))
    series_id_by_dest = {}
    series_ps_by_dest = {}
    for i, cluster in enumerate(near_dup_clusters, start=1):
        for dest in cluster:
            series_id_by_dest[dest] = i
        series_ps_by_dest.update(_series_ps_commands(cluster))

    # PROMPT_report_detail_xlsx.md, "Примечание": «дата приблизительная» -- Tier B/C
    # (дата ЕСТЬ, но не по EXIF, run_logs.date_review()/dates_review.csv, только
    # image/video-ветка _process_record() -- RAW/DVD туда не попадают, lookup для них просто
    # не срабатывает, без отдельного гейта). Tier A (dest отсутствует в dates_review.csv) --
    # доверенная дата, никакой пометки не нужно, тем же принципом, что и tier_counts["A"] в
    # build_model_from_rows().
    approx_date_dests = {row.get("dest") for row in data.get("dates_review", [])
                          if row.get("dest") and row.get("tier") in ("B", "C")}

    rows = []
    # REVIEW-HANDOFF.md, Раунд 95 [БЛОКЕР]: _process_dvd_item() (photosort_win.py:7593-7628)
    # вызывает run_logs.appended() БЕЗУСЛОВНО на КАЖДЫЙ файл DVD-юнита (SourceWalker.
    # _handle_dvd_unit() yield'ит один SourceItem на файл, :3598-3609) -- appended.csv для
    # одного DVD-рипа реально содержит 10-30+ строк с одинаковым reason, не одну. "Одна строка
    # в appended.csv" (докстринг этой функции ДО фикса Раунда 95) было утверждением о коде,
    # непроверенным исполнением -- проверка показала обратное. Группируем по _dvd_unit_root(dest)
    # (Раунд 96 придирка: не просто dirname(dest) -- см. её докстринг выше) --
    # все файлы одного юнита физически лежат в одной VIDEO_TS-папке (report.py делает то же
    # самое различие между "постфактум-реклассификация appended.csv" и "живой реестр" для
    # HTML-версии, см. run_stats["dvd_units_copied"] в _render_run_copied() -- но тот реестр
    # не долетает досюда (generate_detail_xlsx() получает только data_new, не run_stats) и,
    # что важнее, ФИЛЬТРУЕТ частично скопированные юниты целиком (Раунд 71) -- для построчной
    # детализации ЭТОГО прогона честнее показать реально скопированные файлы, сколько бы их ни
    # было, чем молчать о частичном юните вовсе.
    dvd_groups = {}
    for r in data.get("appended", []):
        reason = r.get("reason", "") or ""
        if reason.startswith("DVD-Video"):
            dest = r.get("dest", "") or ""
            dvd_groups.setdefault(_dvd_unit_root(dest), []).append(r)
            continue
        source = r.get("source", "") or ""
        dest = r.get("dest", "") or ""
        source_name = _source_basename(source)
        kind = _row_kind(dest, source)
        final_name = _source_basename(dest) if _source_basename(dest) != source_name else ""
        series_id = series_id_by_dest.get(dest, 0)
        notes = []
        if final_name:
            notes.append("переименовано")
        if series_id:
            notes.append("похожая серия")
        if dest in approx_date_dests:
            notes.append("дата приблизительная")
        note = "; ".join(notes)
        rows.append({
            "folder": _source_dirname(source), "name": source_name, "ext": _ext(source),
            "kind": kind, "copied": True, "dest_or_dup": dest, "final_name": final_name,
            "series_id": series_id, "note": note, "color": None,
            "ps_command": series_ps_by_dest.get(dest, ""),
        })

    for dest_dir, group_rows in dvd_groups.items():
        # Представитель группы: любая строка юнита несёт тот же disp_base (та же физическая
        # DVD-папка) -- _source_dirname() на её "source" в обычном (не вложенном) DVD-рипе
        # даёт ровно disp_base, см. _dvd_unit_file_records()/_handle_dvd_unit() (rel -- как
        # правило голое имя файла, без подпапок).
        sample_source = group_rows[0].get("source", "") or ""
        dest_name = _source_basename(dest_dir)
        rows.append({
            "folder": _source_dirname(sample_source), "name": "VIDEO_TS", "ext": "",
            "kind": "video", "copied": True, "dest_or_dup": dest_dir,
            "final_name": dest_name if dest_name != "VIDEO_TS" else "",
            "series_id": 0,
            "note": f"DVD-видео (VIDEO_TS), скопировано целиком ({_n_files(len(group_rows))})",
            "color": None, "ps_command": "",
        })

    for r in data.get("disputes", []):
        source = r.get("source", "") or ""
        dest = r.get("dest", "") or ""
        reason = r.get("reason", "") or ""
        source_name = _source_basename(source)
        final_name = _source_basename(dest) if dest and _source_basename(dest) != source_name else ""
        notes = [f"спорный (_Unsorted): {_dispute_reason_label(reason)}"]
        if final_name:
            notes.append("переименовано")
        rows.append({
            "folder": _source_dirname(source), "name": source_name, "ext": _ext(source),
            "kind": _row_kind(dest, source), "copied": True, "dest_or_dup": dest,
            "final_name": final_name, "series_id": 0, "note": "; ".join(notes), "color": None,
            "ps_command": "",
        })

    for r in data.get("skipped", []):
        source = r.get("source", "") or ""
        matched = r.get("matched_with", "") or ""
        reason = r.get("reason", "") or ""
        source_name = _source_basename(source)
        if reason == "raw_skipped_has_jpeg":
            kind = "raw"
            note = "RAW не сохранён — есть JPEG-пара, RAW-зеркало отключено"
            color = _COLOR_RAW_SKIPPED
        else:
            # already_present/identical_at_destination -- обе означают "дубликат" (спека).
            kind = _row_kind(matched, source)
            note = "дубликат"
            color = _COLOR_DUPLICATE
        rows.append({
            "folder": _source_dirname(source), "name": source_name, "ext": _ext(source),
            "kind": kind, "copied": False, "dest_or_dup": matched, "final_name": "",
            "series_id": 0, "note": note, "color": color,
            "ps_command": _dup_pair_ps_command(source, matched),
        })

    for r in data.get("unreadable", []):
        source = r.get("source", "") or ""
        error = r.get("error", "") or ""
        source_name = _source_basename(source)
        rows.append({
            "folder": _source_dirname(source), "name": source_name, "ext": _ext(source),
            "kind": _row_kind(source), "copied": False, "dest_or_dup": "", "final_name": "",
            "series_id": 0, "note": f"не прочитано: {error}" if error else "не прочитано",
            "color": _COLOR_PROBLEM, "ps_command": "",
        })

    rows.sort(key=lambda d: (d["folder"], d["name"]))
    return rows


def generate_detail_xlsx(data: dict, report_out_path: str) -> str:
    """Пишет файл-сосед report_out_path (тот же каталог, DETAIL_XLSX_FILENAME) -- тот же
    паттерн размещения, что уже даёт generate_dedup_verification_page() в report.py (см.
    спеку, "Где лежит и как открывается"). Возвращает имя файла (относительный href для
    кнопки «Детализированный отчёт») или None, если строк вообще нет (прогон ничего не
    скопировал/не пропустил/не отклонил) -- тогда файл не пишется и кнопка остаётся
    неактивной, тем же принципом, что и у companion-страницы дублей.

    Открытые решения этой реализации (сознательные, не додуманные молча -- см. CLAUDE.md,
    "Замеченный, но не озвученный побочный эффект"):
    - "Тип медиа" технически может быть "прочее" для спорных файлов без распознаваемого
      расширения (media_note="not_media") -- спека утверждает "ровно три значения", но
      скрывать реальное расширение под неверной меткой (фото/видео/RAW) было бы хуже, чем
      честное "прочее" в редком случае.
    - DVD-юнит -- ОДНА строка на юнит (не на файл, REVIEW-HANDOFF.md, Раунд 95 [БЛОКЕР] --
      appended.csv реально содержит по строке на каждый .VOB/.IFO/.BUP, группировка по
      dirname(dest) внутри этой функции). "Имя файла" -- литерал "VIDEO_TS" (не имя
      конкретного файла), "Итоговое имя файла" -- basename dest-папки, ЕСЛИ отличается от
      "VIDEO_TS" (коллизия дала "VIDEO_TS (2)" и т.п., _unique_dvd_dest_name()) -- та же
      идея "переименовано", что и у обычного файла, просто на уровне папки-юнита.
    - Строки отсортированы по папке, затем по имени файла (_build_detail_rows()) -- так все
      файлы одной папки идут подряд. Раньше поверх этого был Excel-outline (сворачиваемые
      +/- по папке); убран 2026-08-28 вместе с переходом на write_only-режим openpyxl (см.
      _write_flat_xlsx()) -- на большом архиве обычный режим уходил в минуты, а outline для
      прогонов с тысячами папок всё равно бесполезен. Навигация по папке -- автофильтр +
      сортировка по первой колонке.
    - "Открыть файл (PowerShell)" (живая находка пользователя, 2026-09-05: колонка уже была
      в Паспорте, отсутствие в детализации обычного прогона неожиданно для пользователя) --
      две разные логики под одним заголовком, по типу строки: "похожая серия" -- тот же приём,
      что и в Паспорте (_series_ps_commands(), одна команда на весь кластер near_dup, открывает
      всех членов сразу независимо от папки); "дубликат" (skipped: already_present/
      identical_at_destination/raw_skipped_has_jpeg) -- своя логика (_dup_pair_ps_command()):
      ровно одна пара источник (ещё не скопирован, SOURCE) / то, с чем совпал (уже в архиве,
      TARGET) -- команда открывает оба. Остальные строки (обычный новый файл, DVD-юнит,
      спорный, не прочитано) -- пустая ячейка, открывать нечего/не с чем сравнивать."""
    rows = _build_detail_rows(data)
    if not rows:
        return None
    values = [
        [row["folder"], row["name"], row["ext"], _KIND_LABELS.get(row["kind"], row["kind"]),
         "да" if row["copied"] else "нет", row["dest_or_dup"], row["final_name"],
         row["series_id"], row["note"], row["ps_command"]]
        for row in rows
    ]
    out_path = os.path.join(os.path.dirname(report_out_path), DETAIL_XLSX_FILENAME)
    _write_flat_xlsx(_COLUMN_HEADERS, _COLUMN_WIDTHS, values,
                      colors=[row["color"] for row in rows], out_path=out_path,
                      small_font_col=_DETAIL_PS_COMMAND_COL)
    return DETAIL_XLSX_FILENAME


_SMALL_FONT_SIZE = 8  # против дефолтного Calibri 11 -- см. small_font_col ниже


def _write_flat_xlsx(headers: list, widths: list, values: list, colors: list,
                      out_path: str, small_font_col: int = None) -> None:
    """Общая механика записи листа (заголовок жирным, freeze panes, автофильтр, ширины
    колонок, `\\?\\`-безопасное сохранение) -- переиспользуется generate_detail_xlsx() (прогон)
    и generate_passport_detail_xlsx() (Паспорт архива), у которых разная ФОРМА строк
    (9 колонок vs 6), но одинаковая механика листа. `values` -- уже полностью отформатированные
    для отображения списки (никаких bool/Counter/сырых кодов -- ответственность вызывающего
    builder'а). `colors` -- цвет шрифта строки ("#rrggbb") или None; уникальных значений
    единицы, объекты Font кэшируются и один инстанс переиспользуется на все ячейки.

    `small_font_col` (0-based, опционально) -- одна колонка получает уменьшенный шрифт
    (`_SMALL_FONT_SIZE`) независимо от цвета строки. Введён 2026-09-04 для столбца
    "Открыть файл (PowerShell)" в Паспорт-детализации (SESSION-HANDOFF.txt), с 2026-09-05
    передаётся и generate_detail_xlsx() (тот же столбец в детализации обычного прогона) --
    вся ячейка копируется целиком, читаемость на глаз вторична.

    REVIEW-HANDOFF.md, Раунд 204 [ПРИДИРКА] 204-1: `needs_styled_path` ниже гейтит только
    выбор между быстрым (`ws.append`) и медленным (`WriteOnlyCell`) путём -- если строка ушла
    по медленному пути (цветная строка ИЛИ непустая ячейка `small_font_col`), сама ячейка
    `small_font_col` красится мелким шрифтом БЕЗУСЛОВНО, даже когда она пуста (напр. цветная
    строка "дубликат" без пары для сравнения -- `_dup_pair_ps_command()` вернул "").
    Визуально не наблюдаемо (пустая ячейка не рендерит текст ни при каком размере) -- ревизор
    согласился, что править не нужно; имя
    переименовано только для ясности при чтении, не поведение.

    write_only-режим openpyxl (2026-08-28, живой боевой прогон + прямое решение пользователя,
    «вариант 1»): на большом архиве (десятки тысяч спорных/дублей) обычный режим openpyxl
    уходил в 2-3 минуты и сотни МБ RSS на записи ЭТОГО одного файла (создаётся объект Cell на
    каждую ячейку + RowDimension на строку), а окно GUI «Работа окончена» физически ждёт
    возврата -- пользователь решил, что программа зависла. write_only -- ~×35 (170 с -> 4,8 с
    на 40k строк, боевой замер; на другом железе множитель меньше, порядок тот же -- секунды).
    Цена: посгруппная сворачиваемость по папке (Excel outline, `row_dimensions`) в write_only
    не поддерживается и убрана -- для прогонов с тысячами папок она бесполезна, для остальных
    маргинальна; строки по-прежнему отсортированы по папке (файлы одной папки идут подряд),
    навигация -- автофильтр + сортировка по первой колонке."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Детализация")
    ws.freeze_panes = "A2"
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bold_font = Font(bold=True)
    header_cells = [WriteOnlyCell(ws, value=h) for h in headers]
    for cell in header_cells:
        cell.font = bold_font
    ws.append(header_cells)

    font_by_color = {}
    small_font_by_color = {}
    for row_values, color in zip(values, colors, strict=True):
        needs_styled_path = small_font_col is not None and row_values[small_font_col]
        if not color and not needs_styled_path:
            ws.append(row_values)
            continue
        row_font = None
        if color:
            row_font = font_by_color.get(color)
            if row_font is None:
                row_font = font_by_color[color] = Font(color=_argb(color))
        cells = [WriteOnlyCell(ws, value=v) for v in row_values]
        for idx, cell in enumerate(cells):
            if idx == small_font_col:
                small_font = small_font_by_color.get(color)
                if small_font is None:
                    kwargs = {"size": _SMALL_FONT_SIZE}
                    if color:
                        kwargs["color"] = _argb(color)
                    small_font = small_font_by_color[color] = Font(**kwargs)
                cell.font = small_font
            elif row_font is not None:
                cell.font = row_font
        ws.append(cells)

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(values) + 1}"

    # Живая находка (боевой прогон ci/windows_ci_test.py::test_long_path, 2026-08-16):
    # openpyxl.Workbook.save() открывает файл сам (zipfile.ZipFile), без "\\?\"-префикса
    # падает FileNotFoundError на TARGET глубже 260 символов -- тот же случай, что уже
    # решён для report.html в report.py:_write() (_winlong()).
    wb.save(_winlong(out_path))


# ============================================================================
# Паспорт архива (self-scan TARGET) -- PROMPT_report_detail_xlsx.md, "Открыто на момент
# записи", решено и реализовано 2026-08-16. Другая форма строки, чем у прогона выше: одна
# строка = одна НАХОДКА (архив/повреждённый файл/дубликат/похожая серия), не source->dest
# событие -- self-scan ничего никуда не копирует, "куда" не существует как понятие.
# ============================================================================

PASSPORT_DETAIL_XLSX_FILENAME = "passport_detail.xlsx"

_PASSPORT_COLUMN_HEADERS = ["Папка", "Имя", "Тип находки", "№ группы",
                             "Примечание", "Открыть файл (PowerShell)"]
_PASSPORT_COLUMN_WIDTHS = [55, 32, 20, 10, 45, 46]
# 0-based -- см. _DETAIL_PS_COMMAND_COL (ПОСЛЕДНЯЯ колонка, переполнение текста не перекрывает
# "Примечание").
_PASSPORT_PS_COMMAND_COL = 5


def _passport_row(path: str, category: str, group_id: int, note: str, color: str,
                   ps_command: str = "") -> dict:
    return {
        "folder": _source_dirname(path), "name": _source_basename(path),
        "category": category, "group_id": group_id, "note": note, "color": color,
        "ps_command": ps_command,
    }


def _passport_abs_path(rel_path: str, target_path: str = None) -> str:
    """exact_dup_edges/near_dup_edges хранят item.origin_display -- TARGET-относительный путь
    (self_scan: cfg.source=TARGET), в отличие от encrypted_archive_paths/failed_archive_paths/
    disputed_records/unreadable_records, которые уже абсолютны (_analyze_source_abs_path()).
    Живая находка (боевой прогон на синтетическом архиве при реализации Паспорт-детализации,
    2026-08-16): без этой склейки "Папка" для дублей/похожих серий показывала бы относительный
    путь ("Albums\\Album1"), а для архивов/битых файлов -- абсолютный
    ("C:\\...\\TARGET\\Albums\\...") в ОДНОЙ и той же колонке одного и того же листа --
    непоследовательно и мешает читать. Тот же приём склейки, что уже даёт _passport_file_link()
    в report.py. target_path=None (вызов без него) -- путь остаётся относительным, лучше
    показать хоть что-то, чем уронить построитель.

    REVIEW-HANDOFF.md, Раунд 97 [ЗАМЕЧАНИЕ]: os.path.join() здесь -- тот же класс ошибки, что
    уже был находкой Раунда 80 у _analyze_source_abs_path() (photosort_win.py) -- os.path на
    не-Windows раннере (public-репозиторий гоняет tests/ на ubuntu-latest в CI, тот же клон
    ревизора) это posixpath, который склеивает СТЫК через "/" независимо от того, что rel_path
    уже нормализован на "\\" -- даёт смешанный путь. Ручная f-string-склейка (тот же приём, что
    уже применён в _analyze_source_abs_path()) даёт идентичный на Windows результат, но
    остаётся корректной строкой и на POSIX."""
    if not target_path or not rel_path:
        return rel_path
    return target_path.rstrip("\\") + "\\" + rel_path


def _build_passport_detail_rows(stats, target_path: str = None) -> list:
    """Плоский построитель для Паспорта -- два реальных пробела спеки (архивы/повреждённые
    файлы с разбивкой по причине, оба УЖЕ считаются AnalyzeStats безусловно, просто не
    читались _render_passport_integrity() до этой правки) + миграция того, что раньше показывал
    passport_verification.html (точные дубли/похожие серии, теперь БЕЗ обрезки по multi-folder
    -- полный список, той же кластеризацией, _cluster_passport_edges(), что уже использует
    HTML-карточка "Целостность архива" для одних только счётчиков).

    Группа дублей/похожих серий может содержать 3+ файлов (речь пользователя, 2026-08-16) --
    union-find в _cluster_passport_edges()/_cluster_near_dup() уже это поддерживает нативно
    (не только пары), "№ группы" -- общий номер для всех членов, не "с кем совпал" (в отличие
    от run-detail'я, где сравнение всегда 1:1 "новый файл vs то, что уже в архиве")."""
    rows = []

    for path in stats.encrypted_archive_paths:
        rows.append(_passport_row(path, "архив", 0, "запаролен", _COLOR_PROBLEM))
    for path in stats.failed_archive_paths:
        rows.append(_passport_row(path, "архив", 0, "не открылся", _COLOR_PROBLEM))

    # disputed_records/unreadable_records -- {"in_archive": bool, "abs_path": str|None,
    # "display": str} (_analyze_dispute_record(), photosort_win.py) -- abs_path пуст только
    # для файлов ИЗНУТРИ архива (физический файл жил во временной tmp_extract-папке, уже
    # вычищенной к моменту чтения отчёта), display -- "витринная" строка-замена в этом случае.
    for rec in stats.disputed_records:
        path = rec.get("abs_path") or rec.get("display") or ""
        rows.append(_passport_row(path, "повреждённый файл", 0, "содержимое не распознано", None))
    for rec in stats.unreadable_records:
        path = rec.get("abs_path") or rec.get("display") or ""
        rows.append(_passport_row(path, "повреждённый файл", 0, "не читается (ошибка чтения)",
                                   _COLOR_PROBLEM))

    exact_clusters = _cluster_passport_edges(stats.exact_dup_edges)
    for i, cluster in enumerate(exact_clusters, start=1):
        abs_paths = [_passport_abs_path(path, target_path) for path in cluster]
        ps_by_path = _series_ps_commands(abs_paths)
        for path in abs_paths:
            rows.append(_passport_row(path, "дубликат", i, "", _COLOR_DUPLICATE,
                                       ps_by_path.get(path, "")))

    near_clusters = _cluster_passport_edges(stats.near_dup_edges)
    for i, cluster in enumerate(near_clusters, start=1):
        abs_paths = [_passport_abs_path(path, target_path) for path in cluster]
        ps_by_path = _series_ps_commands(abs_paths)
        for path in abs_paths:
            rows.append(_passport_row(path, "похожая серия", i, "", None,
                                       ps_by_path.get(path, "")))

    # Живая находка пользователя, 2026-08-24: "N файлов лежат не внутри альбома/даты"
    # (_render_passport_integrity()) раньше был единственным пунктом карточки "Целостность
    # архива" БЕЗ путей вовсе, даже в детализации -- в отличие от архивов/битых файлов/дублей
    # выше, найти КОНКРЕТНЫЙ файл было невозможно. stats.dump_item_paths -- тот же формат пути
    # (item.origin_display), что и exact_dup_edges/near_dup_edges, см. её докстринг.
    for path in stats.dump_item_paths:
        rows.append(_passport_row(_passport_abs_path(path, target_path), "вне альбома/даты", 0,
                                   "", None))

    rows.sort(key=lambda d: (d["folder"], d["name"]))
    return rows


def generate_passport_detail_xlsx(stats, report_out_path: str, target_path: str = None) -> str:
    """Пишет файл-сосед report_out_path (PASSPORT_DETAIL_XLSX_FILENAME) -- тот же паттерн
    размещения/возврата (имя файла или None, если строк нет), что и generate_detail_xlsx()
    выше. Кнопка «Детализированный отчёт» в конце карточки "Целостность архива"
    (_render_passport_integrity(), report.py) становится активной ссылкой на этот файл.
    target_path -- см. _passport_abs_path()."""
    rows = _build_passport_detail_rows(stats, target_path)
    if not rows:
        return None
    values = [[row["folder"], row["name"], row["category"], row["group_id"], row["note"],
               row["ps_command"]] for row in rows]
    out_path = os.path.join(os.path.dirname(report_out_path), PASSPORT_DETAIL_XLSX_FILENAME)
    _write_flat_xlsx(_PASSPORT_COLUMN_HEADERS, _PASSPORT_COLUMN_WIDTHS, values,
                      colors=[row["color"] for row in rows], out_path=out_path,
                      small_font_col=_PASSPORT_PS_COMMAND_COL)
    return PASSPORT_DETAIL_XLSX_FILENAME
