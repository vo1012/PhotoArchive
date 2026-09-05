"""PROMPT_report_detail_xlsx.md, Фаза 0/1 (2026-08-16): построитель плоских строк
(_build_detail_rows) + запись .xlsx (generate_detail_xlsx) в report_detail_xlsx.py, плюс
проводка кнопки «Детализированный отчёт» через report.generate_report(). Плюс (тем же днём,
"Открыто на момент записи" спеки) -- Паспорт архива: _build_passport_detail_rows()/
generate_passport_detail_xlsx(), через report.generate_passport_report()."""
from types import SimpleNamespace

from openpyxl import load_workbook

import report as r
import report_detail_xlsx as rx
from test_report import _FakeAnalyzeStats


def _appended(source, dest, reason="appended_new", ts="2026-01-01 00:00:01"):
    return {"timestamp": ts, "source": source, "dest": dest, "reason": reason, "flags": "",
            "date": "", "duration": "", "place": "", "camera": ""}


def _skipped(source, matched_with, reason, ts="2026-01-01 00:00:01"):
    return {"timestamp": ts, "source": source, "matched_with": matched_with, "reason": reason}


def _disputed(source, dest, reason, ts="2026-01-01 00:00:01"):
    return {"timestamp": ts, "source": source, "reason": reason, "dest": dest, "was_hidden": 0}


def _unreadable(source, error, ts="2026-01-01 00:00:01"):
    return {"timestamp": ts, "source": source, "error": error}


class TestSeriesPsCommandsLengthCap:
    """REVIEW-HANDOFF.md, Раунд 207 [ЗАМЕЧАНИЕ] 207-1: без ограничения длины гигантский кластер
    (серийная съёмка/burst на большом архиве, ~350-400+ членов при типичной длине пути) даёт
    команду длиннее жёсткого лимита ячейки .xlsx (32767 UTF-16 code units) -- openpyxl/Excel
    молча обрезают значение при реальной записи, обрезка может прийтись на середину
    'литерала...' (нечётное число кавычек) -- вся команда становится невыполнимой в PowerShell
    целиком, не "открылись первые N файлов"."""

    def test_small_cluster_unaffected_by_cap(self):
        """Обычный маленький кластер -- байт-в-байт то же поведение, что и до фикса."""
        paths = [r"D:\SOURCE\A\a.jpg", r"D:\SOURCE\A\b.jpg"]
        cmds = rx._series_ps_commands(paths)
        assert cmds[paths[0]] == cmds[paths[1]] == \
            "Start-Process 'D:\\SOURCE\\A\\a.jpg'; Start-Process 'D:\\SOURCE\\A\\b.jpg'"

    def test_huge_cluster_stays_under_hard_excel_cell_limit(self):
        """32767 -- жёсткий лимит формата .xlsx на длину строки в ячейке (не openpyxl-специфика).
        Команда обязана оставаться короче него -- иначе именно на этой правке ловится живая
        находка ревизора."""
        paths = [rf"D:\SOURCE\LongFolderNameForRealisticDepth\Sub\Sub2\IMG_{i:05d}.JPG"
                  for i in range(500)]
        command = rx._series_ps_commands(paths)[paths[0]]
        assert len(command) < 32767
        assert len(command) <= rx._PS_COMMAND_CHAR_BUDGET

    def test_huge_cluster_command_has_even_quote_count_before_comment(self):
        """Обрезка -- только по ЦЕЛЫМ путям, никогда не рвёт Start-Process '...'-литерал
        посередине. Нечётное число кавычек в исполняемой части означало бы незакрытый
        литерал -- вся строка невыполнима в PowerShell (живая находка ревизора, Раунд 207)."""
        paths = [rf"D:\SOURCE\LongFolderNameForRealisticDepth\Sub\Sub2\IMG_{i:05d}.JPG"
                  for i in range(500)]
        command = rx._series_ps_commands(paths)[paths[0]]
        executable_part = command.split("; #")[0]
        assert executable_part.count("'") % 2 == 0

    def test_huge_cluster_notes_omitted_count_as_trailing_comment(self):
        """Остаток кластера, не поместившийся в бюджет, назван PowerShell-комментарием
        (после "#" до конца строки) -- сама команда остаётся синтаксически рабочей и открывает
        первую часть кластера, пользователь явно видит текстом, что кластер не поместился
        целиком (вместо молчаливого битого обрезка от Excel)."""
        paths = [rf"D:\SOURCE\LongFolderNameForRealisticDepth\Sub\Sub2\IMG_{i:05d}.JPG"
                  for i in range(500)]
        command = rx._series_ps_commands(paths)[paths[0]]
        assert "; #" in command
        assert "не поместились" in command
        n_start_process = command.count("Start-Process")
        assert 0 < n_start_process < len(paths)

    def test_huge_cluster_survives_real_xlsx_round_trip_unchanged(self, tmp_path):
        """Проверка исполнением, не теоретическая: реальная запись через openpyxl + чтение
        обратно не обрезает команду (в отличие от состояния до фикса, где 40998-символьная
        команда после round-trip'а становилась 32767-символьной с нечётным числом кавычек)."""
        paths = [rf"D:\SOURCE\LongFolderNameForRealisticDepth\Sub\Sub2\IMG_{i:05d}.JPG"
                  for i in range(500)]
        command = rx._series_ps_commands(paths)[paths[0]]
        out_path = tmp_path / "roundtrip.xlsx"
        rx._write_flat_xlsx(["PS"], [40], [[command]], colors=[None], out_path=str(out_path))
        stored = load_workbook(str(out_path)).active.cell(row=2, column=1).value
        assert stored == command

    def test_one_oversized_outlier_does_not_discard_the_rest_of_cluster(self):
        """REVIEW-HANDOFF.md, Раунд 208 [ПРИДИРКА] 208-1: один аномально длинный путь в
        алфавитно РАННЕЙ позиции (сортировка "/" раньше "\\") раньше отбрасывал (`break`) ВЕСЬ
        остаток кластера, даже прекрасно помещающийся -- живая находка ревизора, проверено
        исполнением. Пропускать нужно ТОЛЬКО непомещающийся элемент (`continue`), не всё, что
        после него."""
        huge_path = "D:/" + ("x" * 40000) + ".jpg"
        normal_path = r"D:\SOURCE\a.jpg"
        command = rx._series_ps_commands([huge_path, normal_path])[normal_path]
        assert "Start-Process 'D:\\SOURCE\\a.jpg'" in command
        assert huge_path not in command
        assert "и ещё 1 файлов не поместились" in command

    def test_all_members_individually_exceed_budget_falls_back_to_comment_only(self):
        """Настоящий крайний случай (не 208-1) -- КАЖДЫЙ путь по отдельности длиннее бюджета,
        буквально нечего включить -- команда честно схлопывается в чистый комментарий (валидная
        пустая инструкция в PowerShell), не падает и не производит незакрытый литерал."""
        huge_paths = ["D:/" + ("x" * 40000) + f"_{i}.jpg" for i in range(2)]
        command = rx._series_ps_commands(huge_paths)[huge_paths[0]]
        assert command.startswith("#")
        assert "Start-Process" not in command


class TestBuildDetailRows:
    def test_normal_appended_row_has_no_note_and_is_copied(self):
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\SOURCE\Vacation\a.jpg", r"D:\TARGET\Albums\Vacation\a.jpg"),
        ]})
        assert len(rows) == 1
        row = rows[0]
        assert row["folder"] == r"D:\SOURCE\Vacation"
        assert row["name"] == "a.jpg"
        assert row["ext"] == "jpg"
        assert row["kind"] == "image"
        assert row["copied"] is True
        assert row["dest_or_dup"] == r"D:\TARGET\Albums\Vacation\a.jpg"
        assert row["final_name"] == ""
        assert row["series_id"] == 0
        assert row["note"] == ""
        assert row["color"] is None

    def test_renamed_dest_sets_final_name_and_note(self):
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\ByDate\2020\a_1.jpg"),
        ]})
        row = rows[0]
        assert row["final_name"] == "a_1.jpg"
        assert "переименовано" in row["note"]

    def test_near_dup_cluster_sets_series_id_and_note_for_all_members(self):
        data = {
            "appended": [
                _appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg"),
                _appended(r"D:\SOURCE\b.jpg", r"D:\TARGET\b.jpg"),
            ],
            "near_dup_edges": [
                {"timestamp": "2026-01-01 00:00:01", "source": r"D:\SOURCE\b.jpg",
                 "dest": r"D:\TARGET\b.jpg", "matched_dest": r"D:\TARGET\a.jpg",
                 "category": "appended_near_dup", "hamming": "3"},
            ],
        }
        rows = rx._build_detail_rows(data)
        by_name = {row["name"]: row for row in rows}
        assert by_name["a.jpg"]["series_id"] == by_name["b.jpg"]["series_id"] == 1
        assert "похожая серия" in by_name["a.jpg"]["note"]
        assert "похожая серия" in by_name["b.jpg"]["note"]

    def test_near_dup_same_folder_dests_get_identical_powershell_command(self):
        """Столбец "Открыть файл (PowerShell)" в детализации обычного прогона (2026-09-05,
        живая находка пользователя -- была только в Паспорте). Живой отзыв пользователя,
        тем же днём: команда должна открывать ВСЮ серию сразу, а не один файл -- у ВСЕХ
        членов одного кластера одна и та же строка (копипаст с любой строки открывает всю
        серию)."""
        data = {
            "appended": [
                _appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\ByDate\2024\a.jpg"),
                _appended(r"D:\SOURCE\b.jpg", r"D:\TARGET\ByDate\2024\b.jpg"),
            ],
            "near_dup_edges": [
                {"timestamp": "2026-01-01 00:00:01", "source": r"D:\SOURCE\b.jpg",
                 "dest": r"D:\TARGET\ByDate\2024\b.jpg",
                 "matched_dest": r"D:\TARGET\ByDate\2024\a.jpg",
                 "category": "appended_near_dup", "hamming": "3"},
            ],
        }
        rows = rx._build_detail_rows(data)
        by_name = {row["name"]: row for row in rows}
        expected = ("Start-Process 'D:\\TARGET\\ByDate\\2024\\a.jpg'; "
                    "Start-Process 'D:\\TARGET\\ByDate\\2024\\b.jpg'")
        assert by_name["a.jpg"]["ps_command"] == expected
        assert by_name["b.jpg"]["ps_command"] == expected

    def test_near_dup_cross_folder_dests_still_get_same_command(self):
        """Развилка, поставленная пользователем 2026-09-05: серию НЕ дробим по папкам --
        Start-Process открывает каждый файл СВОЕЙ программой (не Проводник), расположение
        файла не мешает визуальному сравнению. Разные папки -- команда всё равно одна на
        обоих, открывает обоих."""
        data = {
            "appended": [
                _appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\Albums\A\a.jpg"),
                _appended(r"D:\SOURCE\b.jpg", r"D:\TARGET\ByDate\2024\b.jpg"),
            ],
            "near_dup_edges": [
                {"timestamp": "2026-01-01 00:00:01", "source": r"D:\SOURCE\b.jpg",
                 "dest": r"D:\TARGET\ByDate\2024\b.jpg",
                 "matched_dest": r"D:\TARGET\Albums\A\a.jpg",
                 "category": "appended_near_dup", "hamming": "3"},
            ],
        }
        rows = rx._build_detail_rows(data)
        by_name = {row["name"]: row for row in rows}
        expected = ("Start-Process 'D:\\TARGET\\Albums\\A\\a.jpg'; "
                    "Start-Process 'D:\\TARGET\\ByDate\\2024\\b.jpg'")
        assert by_name["a.jpg"]["ps_command"] == expected
        assert by_name["b.jpg"]["ps_command"] == expected

    def test_ordinary_appended_row_without_series_has_no_command(self):
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg"),
        ]})
        assert rows[0]["ps_command"] == ""

    def test_tier_b_or_c_date_sets_approximate_date_note(self):
        """PROMPT_report_detail_xlsx.md, "Примечание": «дата приблизительная» -- дата ЕСТЬ, но
        не по EXIF (Tier B/C, dates_review.csv). Tier A (dest отсутствует в dates_review.csv)
        -- без пометки."""
        data = {
            "appended": [
                _appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg"),
                _appended(r"D:\SOURCE\b.jpg", r"D:\TARGET\b.jpg"),
                _appended(r"D:\SOURCE\c.jpg", r"D:\TARGET\c.jpg"),
            ],
            "dates_review": [
                {"timestamp": "2026-01-01 00:00:01", "dest": r"D:\TARGET\a.jpg",
                 "date": "2020-01-01", "tier": "B", "confidence": "", "evidence": "",
                 "source": r"D:\SOURCE\a.jpg"},
                {"timestamp": "2026-01-01 00:00:01", "dest": r"D:\TARGET\b.jpg",
                 "date": "2020-01-01", "tier": "C", "confidence": "", "evidence": "",
                 "source": r"D:\SOURCE\b.jpg"},
            ],
        }
        rows = rx._build_detail_rows(data)
        by_name = {row["name"]: row for row in rows}
        assert "дата приблизительная" in by_name["a.jpg"]["note"]
        assert "дата приблизительная" in by_name["b.jpg"]["note"]
        assert "дата приблизительная" not in by_name["c.jpg"]["note"]
        assert by_name["c.jpg"]["note"] == ""

    def test_dvd_unit_multi_file_appended_collapses_to_one_row(self):
        """REVIEW-HANDOFF.md, Раунд 95 [БЛОКЕР]: _process_dvd_item() (photosort_win.py) вызывает
        run_logs.appended() на КАЖДЫЙ файл юнита -- appended.csv реально содержит одну строку
        на .VOB/.IFO/.BUP, не одну на весь диск. Сценарий ревизора (4 файла одного юнита,
        общий reason/dest-папка) -- построитель обязан свернуть их в одну строку."""
        dvd_reason = "DVD-Video (VIDEO_TS), скопирован целиком"
        dest_dir = r"D:\TARGET\Albums\DVD5\VIDEO_TS"
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\RIP\DVD5/VIDEO_TS.IFO", dest_dir + r"\VIDEO_TS.IFO", reason=dvd_reason),
            _appended(r"D:\RIP\DVD5/VTS_01_0.BUP", dest_dir + r"\VTS_01_0.BUP", reason=dvd_reason),
            _appended(r"D:\RIP\DVD5/VTS_01_0.VOB", dest_dir + r"\VTS_01_0.VOB", reason=dvd_reason),
            _appended(r"D:\RIP\DVD5/VTS_01_1.VOB", dest_dir + r"\VTS_01_1.VOB", reason=dvd_reason),
        ]})
        assert len(rows) == 1
        row = rows[0]
        assert row["name"] == "VIDEO_TS"
        assert row["kind"] == "video"
        assert row["copied"] is True
        assert row["dest_or_dup"] == dest_dir
        assert row["final_name"] == ""
        assert row["series_id"] == 0
        assert "DVD" in row["note"]
        assert "4 файла" in row["note"]
        assert row["ps_command"] == ""

    def test_dvd_unit_dest_collision_suffix_shown_as_final_name(self):
        """_unique_dvd_dest_name() суффиксирует папку при коллизии двух физически разных
        дисков ("VIDEO_TS (2)") -- та же идея "переименовано", что и у обычного файла, но на
        уровне папки-юнита."""
        dvd_reason = "DVD-Video (VIDEO_TS), скопирован целиком"
        dest_dir = r"D:\TARGET\Albums\DVD5\VIDEO_TS (2)"
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\RIP\DVD5/VIDEO_TS.IFO", dest_dir + r"\VIDEO_TS.IFO", reason=dvd_reason),
        ]})
        assert rows[0]["final_name"] == "VIDEO_TS (2)"

    def test_two_distinct_dvd_units_stay_separate_rows(self):
        dvd_reason = "DVD-Video (VIDEO_TS), скопирован целиком"
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\RIP\DiskA/VIDEO_TS.IFO", r"D:\TARGET\A\VIDEO_TS\VIDEO_TS.IFO", reason=dvd_reason),
            _appended(r"D:\RIP\DiskA/VTS_01_0.VOB", r"D:\TARGET\A\VIDEO_TS\VTS_01_0.VOB", reason=dvd_reason),
            _appended(r"D:\RIP\DiskB/VIDEO_TS.IFO", r"D:\TARGET\B\VIDEO_TS\VIDEO_TS.IFO", reason=dvd_reason),
        ]})
        assert len(rows) == 2
        note_by_dest = {row["dest_or_dup"]: row["note"] for row in rows}
        assert "2 файла" in note_by_dest[r"D:\TARGET\A\VIDEO_TS"]
        assert "1 файл)" in note_by_dest[r"D:\TARGET\B\VIDEO_TS"]

    def test_dvd_unit_nested_subfolder_files_still_collapse_to_one_row(self):
        """REVIEW-HANDOFF.md, Раунд 96 (придирка): _dvd_unit_file_records()
        (photosort_win.py) рекурсивна "на случай нестандартного рипа" -- если рип нестандартный
        и часть файлов юнита лежит во вложенной подпапке ВНУТРИ VIDEO_TS (не только плоско в её
        корне), os.path.dirname(dest) для этих файлов раньше давал .../VIDEO_TS/Sub, а не
        .../VIDEO_TS -- юнит раскалывался на 2 строки вместо 1. Group-by теперь ищет ближайшего
        предка с именем VIDEO_TS (_dvd_unit_root()), не просто прямого родителя."""
        dvd_reason = "DVD-Video (VIDEO_TS), скопирован целиком"
        dest_dir = r"D:\TARGET\Albums\DVD5\VIDEO_TS"
        rows = rx._build_detail_rows({"appended": [
            _appended(r"D:\RIP\DVD5/VIDEO_TS.IFO", dest_dir + r"\VIDEO_TS.IFO", reason=dvd_reason),
            _appended(r"D:\RIP\DVD5/Sub/VTS_01_0.VOB", dest_dir + r"\Sub\VTS_01_0.VOB",
                      reason=dvd_reason),
        ]})
        assert len(rows) == 1
        assert rows[0]["dest_or_dup"] == dest_dir
        assert "2 файла" in rows[0]["note"]

    def test_skipped_present_is_duplicate_gray_not_copied(self):
        rows = rx._build_detail_rows({"skipped": [
            _skipped(r"D:\SOURCE\dup.jpg", r"D:\TARGET\a.jpg", "already_present"),
        ]})
        row = rows[0]
        assert row["copied"] is False
        assert row["dest_or_dup"] == r"D:\TARGET\a.jpg"
        assert row["note"] == "дубликат"
        assert row["color"] == rx._COLOR_DUPLICATE

    def test_skipped_duplicate_powershell_command_opens_both_files(self):
        """Дубликат обычного прогона -- не кластер self-scan, а ровно одна пара: новый файл
        (ещё на SOURCE) vs то, с чем он совпал (уже в TARGET). Команда открывает ОБА -- сценарий
        "сравнить дубли", ради которого пользователь спросил про отсутствие этой колонки."""
        rows = rx._build_detail_rows({"skipped": [
            _skipped(r"D:\SOURCE\dup.jpg", r"D:\TARGET\a.jpg", "already_present"),
        ]})
        assert rows[0]["ps_command"] == \
            "Start-Process 'D:\\SOURCE\\dup.jpg'; Start-Process 'D:\\TARGET\\a.jpg'"

    def test_skipped_duplicate_quote_in_either_path_escaped_doubled(self):
        rows = rx._build_detail_rows({"skipped": [
            _skipped(r"D:\SOURCE\O'Brien.jpg", r"D:\TARGET\a.jpg", "already_present"),
        ]})
        assert rows[0]["ps_command"] == \
            "Start-Process 'D:\\SOURCE\\O''Brien.jpg'; Start-Process 'D:\\TARGET\\a.jpg'"

    def test_identical_at_destination_is_also_duplicate(self):
        rows = rx._build_detail_rows({"skipped": [
            _skipped(r"D:\SOURCE\dup.jpg", r"D:\TARGET\a.jpg", "identical_at_destination"),
        ]})
        assert rows[0]["note"] == "дубликат"

    def test_raw_skipped_has_jpeg_own_color_and_wording(self):
        """PROMPT_report_detail_xlsx.md, колонка 6: matched_with уже несёт путь JPEG-партнёра
        В АРХИВЕ (photosort_win.py:_process_record(), decision.decision=="raw_skipped") --
        построитель читает готовое значение, не пересчитывает."""
        rows = rx._build_detail_rows({"skipped": [
            _skipped(r"D:\SOURCE\a.cr2", r"D:\TARGET\Albums\a.jpg", "raw_skipped_has_jpeg"),
        ]})
        row = rows[0]
        assert row["kind"] == "raw"
        assert row["copied"] is False
        assert row["dest_or_dup"] == r"D:\TARGET\Albums\a.jpg"
        assert "RAW не сохранён" in row["note"]
        assert row["color"] == rx._COLOR_RAW_SKIPPED
        assert row["color"] not in (rx._COLOR_DUPLICATE, rx._COLOR_PROBLEM)
        assert row["ps_command"] == \
            "Start-Process 'D:\\SOURCE\\a.cr2'; Start-Process 'D:\\TARGET\\Albums\\a.jpg'"

    def test_disputed_row_is_copied_yes_with_unsorted_note(self):
        rows = rx._build_detail_rows({"disputes": [
            _disputed(r"D:\SOURCE\icon.svg", r"D:\TARGET\_Unsorted\icon.svg", "icon_or_svg"),
        ]})
        row = rows[0]
        assert row["copied"] is True
        assert row["dest_or_dup"] == r"D:\TARGET\_Unsorted\icon.svg"
        assert "спорный (_Unsorted)" in row["note"]
        assert "похоже на иконку" in row["note"]  # _dispute_reason_label(), переиспользован
        assert row["color"] is None  # без отдельного цвета -- решение пользователя 2026-08-15
        assert row["ps_command"] == ""  # не пара "дубликат", открывать не с чем сравнивать

    def test_disputed_row_without_recognized_extension_falls_back_to_other(self):
        rows = rx._build_detail_rows({"disputes": [
            _disputed(r"D:\SOURCE\weird.dat", r"D:\TARGET\_Unsorted\weird.dat", "not_media"),
        ]})
        assert rows[0]["kind"] == "other"

    def test_unreadable_row_not_copied_empty_dest_terracotta(self):
        rows = rx._build_detail_rows({"unreadable": [
            _unreadable(r"D:\SOURCE\broken.jpg", "unsupported_container"),
        ]})
        row = rows[0]
        assert row["copied"] is False
        assert row["dest_or_dup"] == ""
        assert row["note"] == "не прочитано: unsupported_container"
        assert row["color"] == rx._COLOR_PROBLEM
        assert row["ps_command"] == ""

    def test_rows_sorted_by_folder_then_name(self):
        data = {"appended": [
            _appended(r"D:\SOURCE\Z\a.jpg", r"D:\TARGET\a.jpg"),
            _appended(r"D:\SOURCE\A\b.jpg", r"D:\TARGET\b.jpg"),
            _appended(r"D:\SOURCE\A\a.jpg", r"D:\TARGET\c.jpg"),
        ]}
        rows = rx._build_detail_rows(data)
        assert [(r_["folder"], r_["name"]) for r_ in rows] == [
            (r"D:\SOURCE\A", "a.jpg"), (r"D:\SOURCE\A", "b.jpg"), (r"D:\SOURCE\Z", "a.jpg"),
        ]

    def test_source_dirname_handles_archive_member_slash_path(self):
        """origin_display для файлов из архива использует "/" (report.py:_source_basename()
        докстринг) -- _source_dirname() должен понимать тот же формат."""
        assert rx._source_dirname("Foto.zip/подпапка/файл.jpg") == "Foto.zip/подпапка"
        assert rx._source_dirname(r"D:\SOURCE\a.jpg") == r"D:\SOURCE"


class TestGenerateDetailXlsx:
    def test_returns_none_and_writes_nothing_when_no_rows(self, tmp_path):
        out_path = tmp_path / "report.html"
        result = rx.generate_detail_xlsx({}, str(out_path))
        assert result is None
        assert not (tmp_path / rx.DETAIL_XLSX_FILENAME).exists()

    def test_writes_file_next_to_report_with_header_and_autofilter(self, tmp_path):
        out_path = tmp_path / "report.html"
        data = {"appended": [_appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg")]}
        result = rx.generate_detail_xlsx(data, str(out_path))
        assert result == rx.DETAIL_XLSX_FILENAME
        full_path = tmp_path / rx.DETAIL_XLSX_FILENAME
        assert full_path.exists()
        wb = load_workbook(str(full_path))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header == rx._COLUMN_HEADERS
        assert ws[1][0].font.bold is True
        assert ws.auto_filter.ref == f"A1:J{ws.max_row}"
        assert ws.freeze_panes == "A2"

    def test_many_same_color_rows_all_get_the_color(self, tmp_path):
        """Живой боевой прогон 2026-08-28: _write_flat_xlsx() кэширует объекты Font по цвету и
        переиспределяет один инстанс на все ячейки одного цвета (было -- новый Font на каждую
        ячейку, сотни тысяч на большом архиве). Проверяем, что цвет от этого не теряется:
        несколько строк одного цвета -- все реально окрашены."""
        data = {"skipped": [
            _skipped(rf"D:\SOURCE\dup{i}.jpg", rf"D:\TARGET\a{i}.jpg", "already_present")
            for i in range(5)
        ]}
        rx.generate_detail_xlsx(data, str(tmp_path / "report.html"))
        wb = load_workbook(str(tmp_path / rx.DETAIL_XLSX_FILENAME))
        ws = wb.active
        expected = "FF" + rx._COLOR_DUPLICATE.lstrip("#")
        for row_i in range(2, 7):  # 5 строк данных
            assert ws.cell(row=row_i, column=1).font.color.rgb == expected

    def test_rows_grouped_by_folder_in_output_order(self, tmp_path):
        """2026-08-28 (write_only-режим, "вариант 1"): Excel-outline (сворачиваемые +/- по
        папке) убран ради ×40 скорости на большом архиве -- но строки по-прежнему
        отсортированы по папке (файлы одной папки идут подряд), навигация через автофильтр/
        сортировку. Проверяем именно порядок (не outline_level, которого больше нет)."""
        out_path = tmp_path / "report.html"
        data = {"appended": [
            _appended(r"D:\SOURCE\B\a.jpg", r"D:\TARGET\c.jpg"),
            _appended(r"D:\SOURCE\A\b.jpg", r"D:\TARGET\b.jpg"),
            _appended(r"D:\SOURCE\A\a.jpg", r"D:\TARGET\a.jpg"),
        ]}
        rx.generate_detail_xlsx(data, str(out_path))
        wb = load_workbook(str(tmp_path / rx.DETAIL_XLSX_FILENAME))
        ws = wb.active
        folders = [ws.cell(row=i, column=1).value for i in (2, 3, 4)]
        names = [ws.cell(row=i, column=2).value for i in (2, 3, 4)]
        assert folders == [r"D:\SOURCE\A", r"D:\SOURCE\A", r"D:\SOURCE\B"]  # папка A подряд
        assert names == ["a.jpg", "b.jpg", "a.jpg"]  # внутри папки -- по имени
        # write_only больше не пишет row_dimensions -- их нет/дефолтны, это ожидаемо
        assert ws.row_dimensions[3].outline_level == 0

    def test_powershell_column_uses_small_font_when_populated(self, tmp_path):
        data = {"skipped": [
            _skipped(r"D:\SOURCE\dup.jpg", r"D:\TARGET\a.jpg", "already_present"),
        ]}
        rx.generate_detail_xlsx(data, str(tmp_path / "report.html"))
        wb = load_workbook(str(tmp_path / rx.DETAIL_XLSX_FILENAME))
        ws = wb.active
        ps_cell = ws.cell(row=2, column=rx._DETAIL_PS_COMMAND_COL + 1)
        assert ps_cell.value.startswith("Start-Process ")
        assert ps_cell.font.size == rx._SMALL_FONT_SIZE

    def test_duplicate_and_problem_rows_get_distinct_visible_font_colors(self, tmp_path):
        out_path = tmp_path / "report.html"
        data = {
            "skipped": [_skipped(r"D:\SOURCE\dup.jpg", r"D:\TARGET\a.jpg", "already_present")],
            "unreadable": [_unreadable(r"D:\SOURCE\broken.jpg", "err")],
        }
        rx.generate_detail_xlsx(data, str(out_path))
        wb = load_workbook(str(tmp_path / rx.DETAIL_XLSX_FILENAME))
        ws = wb.active
        # broken.jpg (unreadable) сортируется раньше dup.jpg (skipped) -- оба в D:\SOURCE.
        rgb_by_name = {ws.cell(row=i, column=2).value: ws.cell(row=i, column=1).font.color.rgb
                       for i in (2, 3)}
        assert rgb_by_name["broken.jpg"] == "FF" + rx._COLOR_PROBLEM.lstrip("#")
        assert rgb_by_name["dup.jpg"] == "FF" + rx._COLOR_DUPLICATE.lstrip("#")
        assert rgb_by_name["broken.jpg"] != rgb_by_name["dup.jpg"]


class TestGenerateReportWiresDetailXlsxButton:
    def test_target_level_with_new_files_gets_active_link_and_file(self, tmp_path):
        data = {"appended": [_appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg")]}
        out_path = tmp_path / "report.html"
        r.generate_report(data, str(out_path), level="target", run_start="2026-01-01 00:00:00")
        html_out = out_path.read_text(encoding="utf-8")
        assert f'href="{rx.DETAIL_XLSX_FILENAME}"' in html_out
        assert 'disabled aria-disabled="true">Детализированный отчёт' not in html_out
        assert (tmp_path / rx.DETAIL_XLSX_FILENAME).exists()

    def test_workdir_level_also_gets_active_link(self, tmp_path):
        """Спека, "Формат и охват": оба режима (level=="target" И level=="workdir") -- один
        код-путь после унификации dry-run 2026-08-14."""
        data = {"appended": [_appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg")]}
        out_path = tmp_path / "report.html"
        r.generate_report(data, str(out_path), level="workdir", run_start="2026-01-01 00:00:00")
        html_out = out_path.read_text(encoding="utf-8")
        assert f'href="{rx.DETAIL_XLSX_FILENAME}"' in html_out

    def test_stays_disabled_when_nothing_at_all_happened(self, tmp_path):
        out_path = tmp_path / "report.html"
        r.generate_report({}, str(out_path), level="target", run_start="2026-01-01 00:00:00")
        html_out = out_path.read_text(encoding="utf-8")
        assert "disabled" in html_out
        assert f'href="{rx.DETAIL_XLSX_FILENAME}"' not in html_out
        assert not (tmp_path / rx.DETAIL_XLSX_FILENAME).exists()

    def test_analyze_level_never_calls_detail_xlsx_at_all(self, tmp_path):
        """checklist_new остаётся None для level=="analyze" (run_start никогда не передаётся
        этой веткой в проде, см. generate_report_from_analyze_stats()) -- эта ветка не должна
        писать report_detail.xlsx вообще, не только не ссылаться на него."""
        data = {"appended": [_appended(r"D:\SOURCE\a.jpg", r"D:\TARGET\a.jpg")]}
        out_path = tmp_path / "report.html"
        r.generate_report(data, str(out_path), level="target")  # run_start не передан
        assert not (tmp_path / rx.DETAIL_XLSX_FILENAME).exists()

    def test_only_skipped_rows_still_activates_button_even_with_zero_new(self, tmp_path):
        """n_new_total==0 (ранний return в _render_run_copied()) не должен означать "нечего
        детализировать" -- в xlsx всё равно попадут дубли/спорные/непрочитанное этого прогона."""
        data = {"skipped": [_skipped(r"D:\SOURCE\dup.jpg", r"D:\TARGET\a.jpg", "already_present")]}
        out_path = tmp_path / "report.html"
        r.generate_report(data, str(out_path), level="target", run_start="2026-01-01 00:00:00",
                           run_stats={"skipped_present": 1})
        html_out = out_path.read_text(encoding="utf-8")
        assert f'href="{rx.DETAIL_XLSX_FILENAME}"' in html_out
        assert (tmp_path / rx.DETAIL_XLSX_FILENAME).exists()


def _dispute_record(abs_path=None, display="", in_archive=False):
    """Та же форма, что _analyze_dispute_record() в photosort_win.py."""
    return {"in_archive": in_archive, "abs_path": abs_path, "display": display}


def _passport_stats(**overrides):
    base = dict(
        encrypted_archive_paths=[], failed_archive_paths=[],
        disputed_records=[], unreadable_records=[],
        exact_dup_edges=[], near_dup_edges=[], dump_item_paths=[],
    )
    base.update(overrides)
    return SimpleNamespace(**base)


class TestBuildPassportDetailRows:
    def test_encrypted_archive_row(self):
        stats = _passport_stats(encrypted_archive_paths=[r"D:\TARGET\Foto.zip"])
        rows = rx._build_passport_detail_rows(stats)
        assert len(rows) == 1
        row = rows[0]
        assert row["folder"] == r"D:\TARGET"
        assert row["name"] == "Foto.zip"
        assert row["category"] == "архив"
        assert row["note"] == "запаролен"
        assert row["color"] == rx._COLOR_PROBLEM
        assert row["group_id"] == 0

    def test_failed_archive_row(self):
        stats = _passport_stats(failed_archive_paths=[r"D:\TARGET\Broken.rar"])
        row = rx._build_passport_detail_rows(stats)[0]
        assert row["category"] == "архив"
        assert row["note"] == "не открылся"
        assert row["color"] == rx._COLOR_PROBLEM

    def test_disputed_record_uses_abs_path_no_color(self):
        stats = _passport_stats(disputed_records=[_dispute_record(abs_path=r"D:\TARGET\weird.dat")])
        row = rx._build_passport_detail_rows(stats)[0]
        assert row["folder"] == r"D:\TARGET"
        assert row["name"] == "weird.dat"
        assert row["category"] == "повреждённый файл"
        assert row["note"] == "содержимое не распознано"
        assert row["color"] is None

    def test_disputed_record_in_archive_falls_back_to_display(self):
        """abs_path пуст для файлов ИЗНУТРИ архива (физический файл уже вычищен из
        tmp_extract к моменту чтения отчёта) -- display -- витринная замена."""
        stats = _passport_stats(disputed_records=[
            _dispute_record(abs_path=None, display="Foto.zip/DCIM/broken.jpg", in_archive=True),
        ])
        row = rx._build_passport_detail_rows(stats)[0]
        assert row["name"] == "broken.jpg"

    def test_unreadable_record_gets_problem_color(self):
        stats = _passport_stats(unreadable_records=[_dispute_record(abs_path=r"D:\TARGET\corrupt.jpg")])
        row = rx._build_passport_detail_rows(stats)[0]
        assert row["category"] == "повреждённый файл"
        assert row["note"] == "не читается (ошибка чтения)"
        assert row["color"] == rx._COLOR_PROBLEM

    def test_exact_dup_group_of_three_shares_one_group_id(self):
        """Речь пользователя, 2026-08-16: дублей может быть больше 1 -- группа не ограничена
        парой. Union-find в _cluster_passport_edges() уже поддерживает 3+ файлов нативно."""
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy1.jpg"},
            {"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy2.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats)
        assert len(rows) == 3
        assert {row["name"] for row in rows} == {"orig.jpg", "copy1.jpg", "copy2.jpg"}
        group_ids = {row["group_id"] for row in rows}
        assert group_ids == {1}
        assert all(row["category"] == "дубликат" for row in rows)
        assert all(row["color"] == rx._COLOR_DUPLICATE for row in rows)

    def test_dup_paths_made_absolute_with_target_path(self):
        """Живая находка (боевой прогон на синтетическом архиве, 2026-08-16): без target_path
        "Папка" дублей/похожих серий оставалась относительной ("Albums\\A"), а у архивов/
        битых файлов -- уже абсолютной, непоследовательно в одной колонке одного листа."""
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy1.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats, target_path=r"D:\TARGET")
        by_name = {row["name"]: row for row in rows}
        assert by_name["orig.jpg"]["folder"] == r"D:\TARGET\Albums\A"

    def test_near_dup_group_no_color_own_category(self):
        stats = _passport_stats(near_dup_edges=[
            {"dest": "ByDate/2024/a.jpg", "matched_dest": "ByDate/2024/b.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats)
        assert all(row["category"] == "похожая серия" for row in rows)
        assert all(row["color"] is None for row in rows)
        assert all(row["group_id"] == 1 for row in rows)

    def test_two_distinct_exact_dup_groups_get_different_numbers(self):
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/a1.jpg", "matched_dest": "Albums/A/a2.jpg"},
            {"dest": "Albums/B/b1.jpg", "matched_dest": "Albums/B/b2.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats)
        group_by_name = {row["name"]: row["group_id"] for row in rows}
        assert group_by_name["a1.jpg"] == group_by_name["a2.jpg"]
        assert group_by_name["b1.jpg"] == group_by_name["b2.jpg"]
        assert group_by_name["a1.jpg"] != group_by_name["b1.jpg"]

    def test_same_folder_dup_pair_gets_identical_powershell_command(self):
        """SESSION-HANDOFF.txt, 2026-09-04 ("серии: РЕШЕНИЕ ПРИНЯТО") + правка 2026-09-05
        (живой отзыв пользователя): команда открывает ВСЮ серию, у обоих членов одна и та же
        строка (не по одному файлу на member)."""
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats, target_path=r"D:\TARGET")
        by_name = {row["name"]: row for row in rows}
        expected = ("Start-Process 'D:\\TARGET\\Albums\\A\\copy.jpg'; "
                    "Start-Process 'D:\\TARGET\\Albums\\A\\orig.jpg'")
        assert by_name["orig.jpg"]["ps_command"] == expected
        assert by_name["copy.jpg"]["ps_command"] == expected

    def test_cross_folder_cluster_still_gets_one_shared_command(self):
        """Развилка, поставленная пользователем 2026-09-05, решена в пользу НЕ дробить серию по
        папкам: тот же файл в альбоме и в ByDate -- Start-Process открывает КАЖДЫЙ своей
        программой (не Проводник), расположение не мешает сравнить -- оба члена получают одну
        и ту же команду, открывающую обоих."""
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/orig.jpg", "matched_dest": "ByDate/2024/orig.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats)
        assert len(rows) == 2
        commands = {row["ps_command"] for row in rows}
        assert len(commands) == 1  # обе строки несут одну и ту же команду
        command = commands.pop()
        assert command.count("Start-Process") == 2
        assert "Albums" in command and "ByDate" in command

    def test_three_way_cluster_all_members_get_same_full_command(self):
        """Кластер из 3 файлов в 2 разных папках -- ВСЕ трое получают одну и ту же команду,
        открывающую всех троих (не только тех, что делят папку)."""
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy1.jpg"},
            {"dest": "Albums/A/orig.jpg", "matched_dest": "ByDate/2024/copy2.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats)
        by_name = {row["name"]: row["ps_command"] for row in rows}
        assert by_name["orig.jpg"] == by_name["copy1.jpg"] == by_name["copy2.jpg"]
        assert by_name["orig.jpg"].count("Start-Process") == 3

    def test_near_dup_cluster_also_gets_shared_command(self):
        """Решение пользователя: столбец применяется к ОБЕИМ категориям, не только "дубликат"."""
        stats = _passport_stats(near_dup_edges=[
            {"dest": "ByDate/2024/a.jpg", "matched_dest": "ByDate/2024/b.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats)
        commands = {row["ps_command"] for row in rows}
        assert commands == {rows[0]["ps_command"]}
        assert rows[0]["ps_command"] != ""

    def test_single_quote_in_path_escaped_doubled(self):
        """Одинарная кавычка в имени файла ломает PowerShell-литерал '...' без экранирования --
        мелкая деталь, явно отмеченная в SESSION-HANDOFF.txt как легко забываемая."""
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/O'Brien.jpg", "matched_dest": "Albums/A/copy.jpg"},
        ])
        rows = rx._build_passport_detail_rows(stats, target_path=r"D:\TARGET")
        by_name = {row["name"]: row for row in rows}
        expected = ("Start-Process 'D:\\TARGET\\Albums\\A\\O''Brien.jpg'; "
                    "Start-Process 'D:\\TARGET\\Albums\\A\\copy.jpg'")
        assert by_name["O'Brien.jpg"]["ps_command"] == expected
        assert by_name["copy.jpg"]["ps_command"] == expected

    def test_dump_item_paths_get_own_category_no_color_no_group(self):
        """Живая находка пользователя, 2026-08-24: "N файлов лежат не внутри альбома/даты"
        (_render_passport_integrity()) раньше был единственным пунктом карточки без путей в
        детализации вовсе -- stats.dump_item_paths теперь тоже попадает в xlsx, тем же
        принципом, что и архивы/битые файлы/дубли выше."""
        stats = _passport_stats(dump_item_paths=["SomeStray/photo.jpg"])
        rows = rx._build_passport_detail_rows(stats)
        assert len(rows) == 1
        row = rows[0]
        assert row["folder"] == "SomeStray"
        assert row["name"] == "photo.jpg"
        assert row["category"] == "вне альбома/даты"
        assert row["note"] == ""
        assert row["color"] is None
        assert row["group_id"] == 0

    def test_dump_item_paths_made_absolute_with_target_path(self):
        stats = _passport_stats(dump_item_paths=["SomeStray/photo.jpg"])
        rows = rx._build_passport_detail_rows(stats, target_path=r"D:\TARGET")
        assert rows[0]["folder"] == r"D:\TARGET\SomeStray"


class TestGeneratePassportDetailXlsx:
    def test_returns_none_when_nothing_found(self, tmp_path):
        out_path = tmp_path / "passport.html"
        assert rx.generate_passport_detail_xlsx(_passport_stats(), str(out_path)) is None
        assert not (tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME).exists()

    def test_writes_file_with_expected_columns(self, tmp_path):
        stats = _passport_stats(encrypted_archive_paths=[r"D:\TARGET\Foto.zip"])
        out_path = tmp_path / "passport.html"
        result = rx.generate_passport_detail_xlsx(stats, str(out_path))
        assert result == rx.PASSPORT_DETAIL_XLSX_FILENAME
        wb = load_workbook(str(tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header == rx._PASSPORT_COLUMN_HEADERS
        row = [c.value for c in ws[2]]
        # openpyxl округляет "" до None на реальном сохранении/чтении файла (не наблюдалось
        # раньше -- предыдущие проверки этого файла сверяли только строящиеся dict'ы, не
        # реально записанный .xlsx) -- сама ячейка пуста, что и требовалось для "нет команды".
        # "Открыть файл (PowerShell)" -- ПОСЛЕДНЯЯ колонка (2026-09-05, не перекрывает
        # "Примечание" переполнением текста).
        assert row == [r"D:\TARGET", "Foto.zip", "архив", 0, "запаролен", None]

    def test_powershell_column_uses_small_font_when_populated(self, tmp_path):
        stats = _passport_stats(exact_dup_edges=[
            {"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy.jpg"},
        ])
        out_path = tmp_path / "passport.html"
        rx.generate_passport_detail_xlsx(stats, str(out_path), target_path=r"D:\TARGET")
        wb = load_workbook(str(tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME))
        ws = wb.active
        ps_cell = ws.cell(row=2, column=rx._PASSPORT_PS_COMMAND_COL + 1)
        assert ps_cell.value.startswith("Start-Process ")
        assert ps_cell.font.size == rx._SMALL_FONT_SIZE
        # соседняя колонка ("Тип находки", раскрашенная _COLOR_DUPLICATE) сохраняет обычный
        # размер -- маленький шрифт не протекает за пределы своей колонки
        category_cell = ws.cell(row=2, column=3)
        assert category_cell.font.size != rx._SMALL_FONT_SIZE


class TestGeneratePassportReportWiresDetailXlsxButton:
    def test_active_link_when_findings_exist(self, tmp_path):
        stats = _FakeAnalyzeStats()
        stats.encrypted_archive_paths = [r"D:\TARGET\Foto.zip"]
        out_path = tmp_path / "passport.html"
        r.generate_passport_report(stats, str(out_path))
        html_out = out_path.read_text(encoding="utf-8")
        assert f'href="{rx.PASSPORT_DETAIL_XLSX_FILENAME}"' in html_out
        assert (tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME).exists()

    def test_target_path_reaches_xlsx_for_absolute_dup_paths(self, tmp_path):
        stats = _FakeAnalyzeStats()
        stats.exact_dup_edges = [{"dest": "Albums/A/orig.jpg", "matched_dest": "Albums/A/copy.jpg"}]
        out_path = tmp_path / "passport.html"
        r.generate_passport_report(stats, str(out_path), target_path=r"D:\TARGET")
        wb = load_workbook(str(tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME))
        ws = wb.active
        folders = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert all(f.startswith(r"D:\TARGET") for f in folders)

    def test_stays_disabled_when_archive_is_clean(self, tmp_path):
        stats = _FakeAnalyzeStats()
        out_path = tmp_path / "passport.html"
        r.generate_passport_report(stats, str(out_path))
        html_out = out_path.read_text(encoding="utf-8")
        assert 'disabled aria-disabled="true">Детализированный отчёт' in html_out
        assert f'href="{rx.PASSPORT_DETAIL_XLSX_FILENAME}"' not in html_out
        assert not (tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME).exists()

    def test_dump_items_alone_activate_the_button(self, tmp_path):
        """Живая находка пользователя, 2026-08-24: "N файлов лежат не внутри альбома/даты" мог
        быть ЕДИНСТВЕННОЙ находкой паспорта -- до фикса dump_item_paths не попадал в xlsx вовсе,
        значит rows была бы пуста и кнопка осталась бы неактивной, хотя report.html уже
        показывает "N файлов..." -- несоответствие между "что видно" и "что можно раскрыть"."""
        stats = _FakeAnalyzeStats()
        stats.n_dump_items = 1
        stats.dump_item_paths = ["SomeStray/photo.jpg"]
        out_path = tmp_path / "passport.html"
        r.generate_passport_report(stats, str(out_path))
        html_out = out_path.read_text(encoding="utf-8")
        assert f'href="{rx.PASSPORT_DETAIL_XLSX_FILENAME}"' in html_out
        assert (tmp_path / rx.PASSPORT_DETAIL_XLSX_FILENAME).exists()


class TestPassportArchivesAndBrokenAttnText:
    def test_archives_breakdown_mentions_encrypted_and_failed(self):
        text = r._passport_archives_attn(3, 1, 1)
        assert "3 архива" in text
        assert "1 запаролен" in text
        assert "1 не открылся" in text

    def test_archives_no_breakdown_sentence_when_all_plain(self):
        text = r._passport_archives_attn(2, 0, 0)
        assert "Из них" not in text

    def test_broken_breakdown_mentions_unreadable_and_rest(self):
        text = r._passport_broken_attn(5, 2)
        assert "5 файлов" in text
        assert "2 не удалось физически прочитать" in text
        assert "остальные 3" in text

    def test_broken_no_breakdown_when_all_unreadable(self):
        text = r._passport_broken_attn(2, 2)
        assert "остальные" not in text

    def test_broken_no_breakdown_sentence_when_none_unreadable(self):
        text = r._passport_broken_attn(4, 0)
        assert "Из них" not in text
